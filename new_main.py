import pandas as pd
import openpyxl
import webbrowser
from urllib.request import urlopen
from selenium import webdriver
from bs4 import BeautifulSoup
import requests
import urllib.request
from datetime import date
from openpyxl import load_workbook
from sqlalchemy import create_engine
import urllib.request, urllib.error, urllib.parse
import yaml
import time



inp=pd.read_excel('input.xlsx')
out=pd.read_excel('Output.xlsx')


conf = yaml.load(open('login.yml'))
myFbEmail = conf['fb_user']['email']
myFbPassword = conf['fb_user']['password']

driver = webdriver.Chrome()
url=inp['WebSite'][0]
def login(url,usernameId, username, passwordId, password, submit_buttonId):
   driver.get(url)
   driver.find_element_by_name(usernameId).send_keys(username)
   driver.find_element_by_name(passwordId).send_keys(password)
   driver.find_element_by_name(submit_buttonId).click()
login(url, "user[email]", myFbEmail, "user[password]", myFbPassword, "commit")
time.sleep(20)
xx=[]
for i in range(len(inp)):
    a=[]
    ID=inp['SynupId'][i]
    url=inp['WebSite'][i]
    
    driver.get(url)
    time.sleep(12)
    soup= BeautifulSoup(driver.page_source)
        
    
    today = date.today()
    Date_Report=str(today)
    
    b=soup.find('button',attrs={'class':'button'})

    if str(type(b))=="<class 'NoneType'>":
        err=soup.findAll('h2',attrs={'class':'Typography__StyledText-sc-1t7fs6h-0 gJiWyz'})[0]
        if str(type(err))=="<class 'NoneType'>":
            error_name='404 not found'
        else:
            error_name=err.text
            workbook_name = 'Output.xlsx'
            wb = load_workbook(workbook_name)
            sheets = wb.sheetnames
            page = wb[sheets[1]]
            new = [[ID,Date_Report,error_name,url]]
            for info in new:
                page.append(info)
            wb.save(filename=workbook_name)
        num=0
        xx=xx+[num]
    else:
        period_time=b.text[:25]
        prof=soup.findAll('span', attrs={'class':'Typography__StyledText-sc-1t7fs6h-0 gcUuHL'})[0]
        profile_views=prof.text
    
        web=soup.findAll('span', attrs={'class':'Typography__StyledText-sc-1t7fs6h-0 gcUuHL'})[1]
        website_views=web.text

        phone=soup.findAll('span', attrs={'class':'Typography__StyledText-sc-1t7fs6h-0 gcUuHL'})[2]
        phone_calls=phone.text

        direction=soup.findAll('span', attrs={'class':'Typography__StyledText-sc-1t7fs6h-0 gcUuHL'})[3]
        direction_requests=direction.text

        bot=soup.findAll('span', attrs={'class':'Typography__StyledText-sc-1t7fs6h-0 gcUuHL'})[4]
        button_clicks=bot.text

        search=soup.findAll('span', attrs={'class':'Typography__StyledText-sc-1t7fs6h-0 gcUuHL'})[5]
        direct_searches=search.text

        discover=soup.findAll('span', attrs={'class':'Typography__StyledText-sc-1t7fs6h-0 gcUuHL'})[6]
        discovery_searches=discover.text
        workbook_name = 'Output.xlsx'
        wb = load_workbook(workbook_name)
        sheets = wb.sheetnames
        page = wb[sheets[0]]
        new = [[ID,period_time,Date_Report,profile_views,website_views,phone_calls,direction_requests,direct_searches,discovery_searches,url]]
        for info in new:
            page.append(info)
        wb.save(filename=workbook_name)
        num=1
        xx=xx+[num]
driver.close()

new_in = pd.read_excel('input.xlsx')
h=len(new_in)
for i in range(h):
    if xx[i]==1:
        new_in=new_in.drop([i])
new_in.reset_index(drop=True, inplace=True)
new_in.to_excel('input.xlsx',index=False)        